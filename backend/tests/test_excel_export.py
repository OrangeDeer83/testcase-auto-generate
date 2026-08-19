import io

from openpyxl import load_workbook

from app.models.test_case import TestCase, TestStep
from app.services.excel_export import to_excel


def _sample_case() -> TestCase:
    return TestCase(
        name="登入成功",
        module="/登入功能",
        preconditions="使用者已註冊帳號",
        steps=[
            TestStep(step_no=1, description="輸入帳號密碼", expected_result="欄位顯示已輸入內容"),
            TestStep(step_no=2, description="點擊登入按鈕", expected_result="導向首頁"),
        ],
        priority="P0",
        notes="備註內容",
    )


def test_to_excel_writes_header_row() -> None:
    workbook = load_workbook(io.BytesIO(to_excel([_sample_case()])))
    sheet = workbook.active

    assert sheet.title == "Test Cases"
    assert [cell.value for cell in sheet[1]] == [
        "用例名稱",
        "所屬模塊",
        "前置條件",
        "步驟描述",
        "預期結果",
        "用例等級",
        "備註",
    ]


def test_to_excel_writes_case_fields_and_formats_steps() -> None:
    workbook = load_workbook(io.BytesIO(to_excel([_sample_case()])))
    row = [cell.value for cell in workbook.active[2]]

    assert row[0] == "登入成功"
    assert row[1] == "/登入功能"
    assert row[2] == "使用者已註冊帳號"
    assert row[3] == "[1] 輸入帳號密碼\n[2] 點擊登入按鈕"
    assert row[4] == "[1] 欄位顯示已輸入內容\n[2] 導向首頁"
    assert row[5] == "P0"
    assert row[6] == "備註內容"


def test_to_excel_skips_steps_without_expected_result() -> None:
    case = TestCase(
        name="部分步驟沒有預期結果",
        steps=[
            TestStep(step_no=1, description="第一步", expected_result="有結果"),
            TestStep(step_no=2, description="第二步", expected_result=""),
        ],
        priority="P1",
    )

    workbook = load_workbook(io.BytesIO(to_excel([case])))
    expected_results = workbook.active[2][4].value

    assert expected_results == "[1] 有結果"


def test_to_excel_with_no_test_cases_only_has_header_row() -> None:
    workbook = load_workbook(io.BytesIO(to_excel([])))

    assert workbook.active.max_row == 1
