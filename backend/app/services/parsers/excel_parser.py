import io

from openpyxl import load_workbook


def extract_excel_text(content: bytes) -> str:
    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    parts: list[str] = []

    for sheet in workbook.worksheets:
        sheet_rows: list[str] = []
        for row in sheet.iter_rows(values_only=True):
            cells = ["" if cell is None else str(cell).strip() for cell in row]
            if any(cells):
                sheet_rows.append(" | ".join(cells))
        if sheet_rows:
            parts.append(f"[工作表：{sheet.title}]")
            parts.extend(sheet_rows)

    return "\n".join(parts)
