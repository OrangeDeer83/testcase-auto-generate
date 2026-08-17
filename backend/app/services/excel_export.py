import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from app.models.test_case import TestCase

HEADERS = ["用例名稱", "所屬模塊", "前置條件", "步驟描述", "預期結果", "用例等級", "備註"]
COLUMN_WIDTHS = [36, 28, 24, 48, 48, 10, 24]


def _steps_text(case: TestCase) -> str:
    return "\n".join(f"[{step.step_no}] {step.description}" for step in case.steps)


def _expected_results_text(case: TestCase) -> str:
    return "\n".join(
        f"[{step.step_no}] {step.expected_result}" for step in case.steps if step.expected_result
    )


def to_excel(test_cases: list[TestCase]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Test Cases"

    sheet.append(HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for case in test_cases:
        sheet.append(
            [
                case.name,
                case.module,
                case.preconditions,
                _steps_text(case),
                _expected_results_text(case),
                case.priority,
                case.notes,
            ]
        )

    wrap = Alignment(wrap_text=True, vertical="top")
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        for cell in row:
            cell.alignment = wrap

    for idx, width in enumerate(COLUMN_WIDTHS, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=idx).column_letter].width = width

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
